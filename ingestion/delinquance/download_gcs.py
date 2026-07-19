"""
ingestion/delinquance/download_gcs.py
=======================================
Télécharge les statistiques de délinquance (SSMSI / Ministère de l'Intérieur)
depuis data.gouv.fr et enrichit communes_agregat avec des indicateurs de sécurité.

Source : data.gouv.fr — "Crimes et délits enregistrés par les forces de sécurité
intérieure depuis 2012" (données annuelles par département).
Dataset ID : 53699576a3a729239d20471d

Stratégie :
  - Les données sont au niveau DÉPARTEMENT (pas commune)
  - On mappe les taux département → toutes les communes du département
  - On garde les 2 dernières années disponibles et on fait la moyenne

Indicateurs calculés par commune (via son département) :
  - taux_cambriolages : cambriolages pour 1 000 logements
  - taux_vols_violence : coups et blessures volontaires pour 1 000 habitants
  - score_securite : score 0-100 (100 = très sûr, inversé des taux)

Usage :
    python3 ingestion/delinquance/download_gcs.py
"""

import io
import re
import requests
import numpy as np
import pandas as pd
import psycopg2
from psycopg2.extras import execute_values
import os

PG_HOST     = os.getenv("POSTGRES_HOST", "localhost")
PG_PORT     = int(os.getenv("POSTGRES_PORT", 5433))
PG_DB       = os.getenv("POSTGRES_DB", "homepedia")
PG_USER     = os.getenv("POSTGRES_USER", "homepedia")
PG_PASSWORD = os.getenv("POSTGRES_PASSWORD", "homepedia")

DEPTS_IDF = {"75", "77", "78", "91", "92", "93", "94", "95"}

# Dataset SSMSI sur data.gouv.fr
DATASET_API = "https://www.data.gouv.fr/api/1/datasets/53699576a3a729239d20471d/"

# Index SSMSI des catégories qui nous intéressent
# Source : notice méthodologique SSMSI
CAMBRIOLAGES_INDEXES = {
    "Cambriolages de logement",
    "Cambriolages d'autres lieux",
}
VOLS_VIOLENCE_INDEXES = {
    "Coups et blessures volontaires sur personnes de 15 ans et plus",
    "Coups et blessures volontaires",
    "Coups et blessures volontaires intrafamiliaux",
    "Violences physiques non crapuleuses",
}


# ─── Helpers ──────────────────────────────────────────────────────────────────

def normalize_dept_code(raw: str) -> str:
    """Normalise un code département ('75' → '75', '092' → '92', '2A' → '2A')."""
    raw = str(raw).strip().upper()
    # Supprimer les zéros à gauche sauf pour les DOM (97x)
    if raw.startswith("0") and not raw.startswith("97"):
        raw = raw.lstrip("0") or "0"
    return raw


def minmax_inverse(series: pd.Series, pct: tuple = (5, 95)) -> pd.Series:
    """Normalise 0-100 inversé (valeur haute → score bas = peu sûr)."""
    low  = np.nanpercentile(series, pct[0])
    high = np.nanpercentile(series, pct[1])
    if high == low:
        return pd.Series(50.0, index=series.index)
    clipped = series.clip(lower=low, upper=high)
    normalized = 100.0 - (clipped - low) / (high - low) * 100.0
    return normalized.round(1)


# ─── Téléchargement données ────────────────────────────────────────────────────

def fetch_dataset_resources() -> list[dict]:
    """Récupère la liste des ressources du dataset data.gouv.fr."""
    r = requests.get(DATASET_API, timeout=30)
    r.raise_for_status()
    data = r.json()
    return data.get("resources", [])


def find_best_csv_resource(resources: list[dict]) -> str | None:
    """Trouve le fichier CSV/XLSX le plus récent contenant des données nationales."""
    candidates = []
    for res in resources:
        title = (res.get("title") or "").lower()
        url   = res.get("url") or ""
        fmt   = (res.get("format") or "").lower()

        # On cherche les fichiers avec données annuelles (pas les mensuels par feuille)
        if any(kw in title for kw in ["annuel", "annual", "département", "national"]):
            candidates.append((res.get("created_at", ""), url, fmt, title))

    # Si aucun candidat ciblé, on prend tous les CSV
    if not candidates:
        for res in resources:
            fmt = (res.get("format") or "").lower()
            url = res.get("url") or ""
            if fmt in ("csv", "xlsx", "xls") and url:
                candidates.append((res.get("created_at", ""), url, fmt,
                                   (res.get("title") or "")))

    candidates.sort(key=lambda x: x[0], reverse=True)
    return (candidates[0][1], candidates[0][2]) if candidates else (None, None)


def download_ssmsi_data() -> pd.DataFrame | None:
    """
    Télécharge et parse les données SSMSI.
    Retourne un DataFrame avec colonnes : code_dept, index_infraction, nb_faits, annee
    """
    print("  ↓ Récupération des ressources data.gouv.fr...")
    try:
        resources = fetch_dataset_resources()
    except Exception as e:
        print(f"  ⚠️  API data.gouv.fr inaccessible : {e}")
        return None

    print(f"     {len(resources)} ressources trouvées")

    # Chercher des fichiers CSV annuels
    csv_resources = [
        r for r in resources
        if (r.get("format") or "").lower() in ("csv",)
        and r.get("url")
    ]
    xlsx_resources = [
        r for r in resources
        if (r.get("format") or "").lower() in ("xlsx", "xls")
        and r.get("url")
    ]

    all_dfs = []

    # 1. Essayer les CSV d'abord (plus faciles à parser)
    for res in sorted(csv_resources, key=lambda r: r.get("created_at", ""), reverse=True)[:3]:
        url   = res["url"]
        title = res.get("title", "")
        try:
            print(f"  ↓ Téléchargement CSV : {title[:60]}...")
            r = requests.get(url, timeout=120)
            r.raise_for_status()
            # Essayer différents séparateurs
            for sep in [";", ",", "\t"]:
                try:
                    df = pd.read_csv(io.BytesIO(r.content), sep=sep, dtype=str,
                                     encoding="utf-8", on_bad_lines="skip")
                    if len(df.columns) >= 4:
                        all_dfs.append(df)
                        print(f"     ✅ {len(df):,} lignes, {len(df.columns)} colonnes")
                        break
                except Exception:
                    continue
        except Exception as e:
            print(f"     ⚠️  Échec : {e}")

    if all_dfs:
        return parse_csv_ssmsi(all_dfs[0])

    # 2. Fallback XLSX
    for res in sorted(xlsx_resources, key=lambda r: r.get("created_at", ""), reverse=True)[:2]:
        url   = res["url"]
        title = res.get("title", "")
        try:
            print(f"  ↓ Téléchargement XLSX : {title[:60]}...")
            r = requests.get(url, timeout=180)
            r.raise_for_status()
            df = parse_xlsx_ssmsi(r.content)
            if df is not None and not df.empty:
                return df
        except Exception as e:
            print(f"     ⚠️  Échec XLSX : {e}")

    return None


def parse_csv_ssmsi(df: pd.DataFrame) -> pd.DataFrame | None:
    """
    Parse un CSV SSMSI au format :
    annee | Code.département | indicateur/libellé | valeur | POP | ...
    Retourne DataFrame: code_dept, libelle, nb_faits, population, annee
    """
    df.columns = [c.strip().lower().replace(".", "_").replace(" ", "_")
                  for c in df.columns]
    print(f"  Colonnes CSV détectées : {list(df.columns)[:10]}")

    # Chercher colonnes clés (fuzzy)
    col_dept  = _find_col(df, ["code_département", "code_departement", "num_dep",
                                "departement", "dept", "dep"])
    col_label = _find_col(df, ["libellé", "libelle", "indicateur", "index",
                                "type_infraction", "categorie"])
    col_val   = _find_col(df, ["valeur", "nombre", "nb_faits", "faits", "count"])
    col_pop   = _find_col(df, ["pop", "population"])
    col_year  = _find_col(df, ["annee", "année", "year"])

    if not col_dept or not col_val:
        print(f"  ⚠️  Colonnes introuvables (dept={col_dept}, val={col_val})")
        return None

    result = pd.DataFrame({
        "code_dept": df[col_dept].astype(str).str.strip(),
        "libelle":   df[col_label].astype(str).str.strip() if col_label else "inconnu",
        "nb_faits":  pd.to_numeric(df[col_val], errors="coerce"),
        "population": pd.to_numeric(df[col_pop], errors="coerce") if col_pop else None,
        "annee":     pd.to_numeric(df[col_year], errors="coerce") if col_year else 2023,
    })
    result["code_dept"] = result["code_dept"].apply(normalize_dept_code)
    result = result[result["code_dept"].isin(DEPTS_IDF)].copy()
    print(f"  → {len(result):,} lignes IDF dans le CSV")
    return result


def parse_xlsx_ssmsi(content: bytes) -> pd.DataFrame | None:
    """
    Parse un XLSX SSMSI (format classique : feuilles = mois, lignes = index crime,
    colonnes = départements).
    Retourne DataFrame: code_dept, libelle, nb_faits, annee (agrégé annuel)
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except ImportError:
        print("  ⚠️  openpyxl non disponible — pip install openpyxl")
        return None

    all_sheets = []
    for sheet_name in wb.sheetnames[:12]:  # max 12 feuilles (mois)
        ws = wb[sheet_name]

        # Récupérer l'année depuis le nom de la feuille (ex: "janv-2023")
        year_match = re.search(r"(\d{4})", sheet_name)
        annee = int(year_match.group(1)) if year_match else 2023

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Première ligne = entêtes (nom de départements ou codes)
        header = [str(c or "").strip() for c in rows[0]]

        # Identifier les colonnes pour l'IDF (noms ou codes 75, 77, 78, 91, 92, 93, 94, 95)
        dept_names_idf = {
            "paris": "75", "seine-et-marne": "77", "yvelines": "78",
            "essonne": "91", "hauts-de-seine": "92",
            "seine-saint-denis": "93", "val-de-marne": "94",
            "val-d'oise": "95", "val d'oise": "95",
        }
        idf_col_indices = {}
        for i, h in enumerate(header):
            h_lower = h.lower().strip()
            for dept_name, dept_code in dept_names_idf.items():
                if dept_name in h_lower:
                    idf_col_indices[i] = dept_code
            # Code numérique direct
            if h_lower in DEPTS_IDF:
                idf_col_indices[i] = h_lower

        if not idf_col_indices:
            continue

        for row in rows[1:]:
            if not row or not row[0]:
                continue
            libelle = str(row[0]).strip()
            if not libelle or libelle.startswith("#"):
                continue
            for col_idx, dept_code in idf_col_indices.items():
                if col_idx < len(row):
                    val = row[col_idx]
                    try:
                        nb = float(str(val).replace(" ", "").replace(",", "."))
                    except (ValueError, TypeError):
                        nb = None
                    if nb is not None:
                        all_sheets.append({
                            "code_dept": dept_code,
                            "libelle": libelle,
                            "nb_faits": nb,
                            "annee": annee,
                        })

    wb.close()

    if not all_sheets:
        return None

    df = pd.DataFrame(all_sheets)
    # Agréger par an (somme des 12 mois)
    df = df.groupby(["code_dept", "libelle", "annee"])["nb_faits"].sum().reset_index()
    print(f"  → {len(df):,} lignes XLSX IDF (après agrégation annuelle)")
    return df


def _find_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    """Cherche le nom de colonne dans df parmi les candidates (case-insensitive)."""
    cols_lower = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in cols_lower:
            return cols_lower[cand.lower()]
    return None


# ─── Calcul des taux par département ──────────────────────────────────────────

def compute_dept_rates(df: pd.DataFrame, conn) -> pd.DataFrame:
    """
    Calcule les taux de délinquance par département IDF.
    Retourne DataFrame: code_dept, taux_cambriolages, taux_vols_violence
    """
    # Populations départements IDF (source INSEE 2022)
    DEPT_POP_IDF = {
        "75": 2_133_111,
        "77": 1_441_305,
        "78": 1_456_299,
        "91": 1_303_639,
        "92": 1_611_662,
        "93": 1_654_994,
        "94": 1_387_926,
        "95": 1_232_939,
    }
    # Nb logements approximatif IDF (source INSEE)
    DEPT_LOG_IDF = {
        "75": 1_408_000,
        "77": 626_000,
        "78": 636_000,
        "91": 576_000,
        "92": 766_000,
        "93": 651_000,
        "94": 616_000,
        "95": 543_000,
    }

    if df is None or df.empty:
        # Données de secours (moyennes IDF 2022 — source ONDRP)
        print("  ⚠️  Données téléchargées vides → utilisation des taux IDF 2022 intégrés")
        fallback_data = {
            "75": (8.2,  7.1),   # (cambriolages/1000 log, CBV/1000 hab)
            "77": (5.8,  4.2),
            "78": (4.7,  3.9),
            "91": (5.1,  4.5),
            "92": (5.4,  4.8),
            "93": (9.3, 10.2),
            "94": (6.3,  5.6),
            "95": (6.8,  5.9),
        }
        rows = [{"code_dept": d, "taux_cambriolages": v[0], "taux_vols_violence": v[1]}
                for d, v in fallback_data.items()]
        return pd.DataFrame(rows)

    # Garder les 2 années les plus récentes
    recent_years = sorted(df["annee"].dropna().unique())[-2:]
    df = df[df["annee"].isin(recent_years)].copy()

    # Classifier les infractions
    def classify(libelle: str) -> str | None:
        lib = libelle.lower()
        if any(kw in lib for kw in ["cambrio", "effract", "burglar"]):
            return "cambriolage"
        if any(kw in lib for kw in ["coups et bless", "violence physiq",
                                     "violence volont", "agression", "cbv"]):
            return "vols_violence"
        return None

    df["categorie"] = df["libelle"].apply(classify)
    df = df[df["categorie"].notna()]

    # Agréger par département + catégorie (moyenne sur les années)
    agg = df.groupby(["code_dept", "categorie"])["nb_faits"].mean().reset_index()
    pivot = agg.pivot(index="code_dept", columns="categorie", values="nb_faits").reset_index()
    pivot.columns.name = None

    # Calculer les taux
    result_rows = []
    for dept in DEPTS_IDF:
        row = pivot[pivot["code_dept"] == dept]
        pop = DEPT_POP_IDF.get(dept, 1_000_000)
        log = DEPT_LOG_IDF.get(dept, 500_000)

        cambrio_raw = float(row["cambriolage"].iloc[0]) if "cambriolage" in row.columns and not row.empty else None
        violence_raw = float(row["vols_violence"].iloc[0]) if "vols_violence" in row.columns and not row.empty else None

        taux_cambrio = round(cambrio_raw / log * 1000, 2) if cambrio_raw else None
        taux_viol    = round(violence_raw / pop * 1000, 2) if violence_raw else None

        result_rows.append({
            "code_dept": dept,
            "taux_cambriolages":  taux_cambrio,
            "taux_vols_violence": taux_viol,
        })

    dept_df = pd.DataFrame(result_rows)
    print("  Taux calculés par département :")
    print(dept_df.to_string(index=False))
    return dept_df


# ─── Score Sécurité ───────────────────────────────────────────────────────────

def compute_score_securite(dept_df: pd.DataFrame) -> pd.DataFrame:
    """
    Score sécurité 0-100 : 100 = très sûr (faibles taux).
    Combine cambriolages (60%) + vols avec violence (40%).

    Normalisation par bornes absolues nationales (source ONDRP/SSMSI 2022)
    plutôt que min-max relatif sur les 8 depts IDF → évite 0 et 100 exacts.
      - Cambriolages : 0‰ → 100pts | 20‰ → 0pts (seuil "zone très à risque")
      - Violence     : 0‰ → 100pts | 25‰ → 0pts
    """
    dept_df = dept_df.copy()

    # Remplir NaN par médiane IDF
    for col in ["taux_cambriolages", "taux_vols_violence"]:
        med = dept_df[col].median()
        dept_df[col] = dept_df[col].fillna(med)

    # Bornes absolues : [seuil_bas (excellent), seuil_haut (très dangereux)]
    # Basées sur les statistiques nationales France métropolitaine
    CAMBRIO_MAX = 20.0   # ‰ logements — au-delà : zone très sensible
    VIOLENCE_MAX = 25.0  # ‰ habitants — au-delà : niveau hors-norme

    cambrio_score = (
        (1.0 - (dept_df["taux_cambriolages"].clip(0, CAMBRIO_MAX) / CAMBRIO_MAX)) * 100
    ).round(1)
    violence_score = (
        (1.0 - (dept_df["taux_vols_violence"].clip(0, VIOLENCE_MAX) / VIOLENCE_MAX)) * 100
    ).round(1)

    dept_df["score_securite"] = (
        cambrio_score * 0.60 + violence_score * 0.40
    ).round(1)

    return dept_df


# ─── Mapping département → communes ──────────────────────────────────────────

def get_communes_by_dept(conn) -> pd.DataFrame:
    """Récupère toutes les communes IDF avec leur département."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT code_commune, TRIM(code_departement) AS code_dept, population_totale
            FROM communes_agregat
            WHERE TRIM(code_departement) = ANY(%s)
        """, (list(DEPTS_IDF),))
        rows = cur.fetchall()
    return pd.DataFrame(rows, columns=["code_commune", "code_dept", "population_totale"])


# ─── PostgreSQL ───────────────────────────────────────────────────────────────

def add_delinquance_columns(conn):
    with conn.cursor() as cur:
        cur.execute("""
            ALTER TABLE communes_agregat
                ADD COLUMN IF NOT EXISTS taux_cambriolages  DOUBLE PRECISION,
                ADD COLUMN IF NOT EXISTS taux_vols_violence DOUBLE PRECISION,
                ADD COLUMN IF NOT EXISTS score_securite     DOUBLE PRECISION
        """)
        conn.commit()
    print("  ✅ Colonnes sécurité prêtes")


def update_delinquance(conn, communes_df: pd.DataFrame, dept_df: pd.DataFrame) -> int:
    """Mappe les taux département aux communes et met à jour la base."""
    merged = communes_df.merge(dept_df, on="code_dept", how="left")

    with conn.cursor() as cur:
        cur.execute("""
            CREATE TEMP TABLE tmp_delinquance (
                code_commune        CHAR(5),
                taux_cambriolages   DOUBLE PRECISION,
                taux_vols_violence  DOUBLE PRECISION,
                score_securite      DOUBLE PRECISION
            ) ON COMMIT DROP
        """)
        data = [
            (
                str(row.code_commune).zfill(5),
                float(row.taux_cambriolages) if pd.notna(row.taux_cambriolages) else None,
                float(row.taux_vols_violence) if pd.notna(row.taux_vols_violence) else None,
                float(row.score_securite) if pd.notna(row.score_securite) else None,
            )
            for row in merged.itertuples()
        ]
        execute_values(cur, "INSERT INTO tmp_delinquance VALUES %s", data)
        cur.execute("""
            UPDATE communes_agregat ca
            SET
                taux_cambriolages   = t.taux_cambriolages,
                taux_vols_violence  = t.taux_vols_violence,
                score_securite      = t.score_securite
            FROM tmp_delinquance t
            WHERE ca.code_commune = t.code_commune
        """)
        updated = cur.rowcount
        conn.commit()
    return updated


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("🔒 Ingestion Délinquance SSMSI → PostgreSQL\n")

    conn = psycopg2.connect(
        host=PG_HOST, port=PG_PORT,
        dbname=PG_DB, user=PG_USER, password=PG_PASSWORD
    )

    # 1. Téléchargement
    df_raw = download_ssmsi_data()

    # 2. Calcul des taux par département
    dept_df = compute_dept_rates(df_raw, conn)

    # 3. Score sécurité
    dept_df = compute_score_securite(dept_df)

    # 4. Ajout colonnes BDD
    add_delinquance_columns(conn)

    # 5. Récupérer les communes
    communes_df = get_communes_by_dept(conn)
    print(f"\n  {len(communes_df):,} communes IDF récupérées")

    # 6. Mise à jour
    updated = update_delinquance(conn, communes_df, dept_df)
    print(f"  → {updated} communes mises à jour ✅\n")

    # 7. Résumé
    print("─" * 65)
    print("Scores sécurité par département IDF :")
    print(f"  {'Dept':<5} {'Cambrio/1k log':>16} {'CBV/1k hab':>12} {'Score':>7}")
    print("  " + "-" * 45)
    for _, r in dept_df.sort_values("score_securite", ascending=False).iterrows():
        dept_name = {
            "75": "Paris", "77": "Seine-et-Marne", "78": "Yvelines",
            "91": "Essonne", "92": "Hauts-de-Seine", "93": "Seine-St-Denis",
            "94": "Val-de-Marne", "95": "Val-d'Oise",
        }.get(r["code_dept"], r["code_dept"])
        print(f"  {r['code_dept']:<5} {(r['taux_cambriolages'] or 0):>14.1f}‰ "
              f"{(r['taux_vols_violence'] or 0):>10.1f}‰ "
              f"{(r['score_securite'] or 0):>6.1f}/100  {dept_name}")

    # Top/Bottom communes
    with conn.cursor() as cur:
        cur.execute("""
            SELECT city, TRIM(code_departement) as dept, score_securite
            FROM communes_agregat
            WHERE score_securite IS NOT NULL
            ORDER BY score_securite DESC
            LIMIT 5
        """)
        top = cur.fetchall()
        cur.execute("""
            SELECT city, TRIM(code_departement) as dept, score_securite
            FROM communes_agregat
            WHERE score_securite IS NOT NULL
            ORDER BY score_securite ASC
            LIMIT 5
        """)
        bottom = cur.fetchall()

    print("\nTop 5 communes les plus sûres (IDF) :")
    for city, dept, score in top:
        print(f"  {(city or '?'):<35} [{dept}]  {score:.1f}/100")
    print("\nBottom 5 communes les moins sûres (IDF) :")
    for city, dept, score in bottom:
        print(f"  {(city or '?'):<35} [{dept}]  {score:.1f}/100")

    conn.close()
    print("\n✅ Ingestion délinquance terminée !")


if __name__ == "__main__":
    main()
